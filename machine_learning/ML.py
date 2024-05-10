import openpyxl
import random
from datetime import datetime

# Đọc file Excel
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

# Thêm cột LED
ws.insert_cols(4)  # Đúng
ws.cell(row=1, column=4).value = "Led"  # Đặt tiêu đề cột
ws.insert_cols(5)  # Đúng
ws.cell(row=1, column=5).value = "Fan"  # Đặt tiêu đề cột
randomRightValue = 0
randomWrongValue = 0
randomRightFan = 0
randomWrongFan = 0
length = len(list(ws.rows))
# Duyệt qua các hàng dữ liệu
for row in ws.iter_rows(min_row=2):  # Bỏ qua hàng tiêu đề
    time_value = row[0].value  # Lấy giá trị thời gian
    date_time = datetime.strptime(time_value, "%d-%m-%Y %H:%M")
    temp_value = row[1].value  # Lấy giá trị nhiệt độ
    humi_value = row[2].value  # Lấy giá trị độ ẩm

    # Xác định giá trị LED
    if date_time.hour in range(6, 10) or date_time.hour in range(18, 22): #6->9, 18->21
        if random.randint(0, 1)  and (1.0*randomRightValue)/length < 0.1:  # 85% có LED = 1
            led_value = 0
            randomRightValue += 1
        else:
            led_value = 1
    else:
        if random.randint(0, 1)  and (1.0*randomWrongValue)/length < 0.1:  # 85% có LED = 1
            led_value = 1
            randomWrongValue += 1
        else:
            led_value = 0
        print(humi_value)
        # Xác định giá trị FAN
    if temp_value > 27.0 or temp_value > 60.0: #6->9, 18->21
        if random.randint(0, 1)  and (1.0*randomRightFan)/length < 0.1:  # 85% có LED = 1
            fan_value = 0
            randomRightFan += 1
        else:
            fan_value = 1
    else:
        if random.randint(0, 1)  and (1.0*randomWrongFan)/length < 0.1:  # 85% có LED = 1
            fan_value = 1
            randomWrongFan += 1
        else:
            fan_value = 0

    # Ghi giá trị LED vào cột
    row[3].value = led_value
    row[4].value = fan_value

# Lưu file Excel đã cập nhật
wb.save('data_with_led.xlsx')